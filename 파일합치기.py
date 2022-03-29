import pandas as pd
import openpyxl

####################################################
Month = '03'
YMD = '20220321'
####################################################

# 엑셀 합치기
fPath = f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}'

Name_A = f"2022년+{Month}월+자체점검실시내역{YMD}"
Name_B = f"2022년+{Month}월+자체점검실시내역{YMD} (1)"
Name_C = f"2022년+{Month}월+자체점검실시내역{YMD} (2)"
Name_D = f"2022년+{Month}월+자체점검실시내역{YMD} (3)"
Name_E = f"2022년+{Month}월+자체점검실시내역{YMD} (4)"
Name_F = f"2022년+{Month}월+자체점검실시내역{YMD} (5)"
Name_G = f"2022년+{Month}월+자체점검실시내역{YMD} (6)"
Name_H = f"2022년+{Month}월+자체점검실시내역{YMD} (7)"
Name_I = f"2022년+{Month}월+자체점검실시내역{YMD} (8)"
Name_J = f"2022년+{Month}월+자체점검실시내역{YMD} (9)"
Name_K = f"2022년+{Month}월+자체점검실시내역{YMD} (10)"
Name_L = f"2022년+{Month}월+자체점검실시내역{YMD} (11)"
Name_M = f"2022년+{Month}월+자체점검실시내역{YMD} (12)"
Name_N = f"2022년+{Month}월+자체점검실시내역{YMD} (13)"
Name_O = f"2022년+{Month}월+자체점검실시내역{YMD} (14)"
Name_P = f"2022년+{Month}월+자체점검실시내역{YMD} (15)"
Name_Q = f"2022년+{Month}월+자체점검실시내역{YMD} (16)"

Name_A = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_A}.xls')
NewName_A = Name_A.drop([0, 1, 2, 3])

Name_B = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_B}.xls')
NewName_B = Name_B.drop([0, 1, 2, 3])

Name_C = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_C}.xls')
NewName_C = Name_C.drop([0, 1, 2, 3])

Name_D = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_D}.xls')
NewName_D = Name_D.drop([0, 1, 2, 3])

Name_E = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_E}.xls')
NewName_E = Name_E.drop([0, 1, 2, 3])

Name_F = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_F}.xls')
NewName_F = Name_F.drop([0, 1, 2, 3])

Name_G = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_G}.xls')
NewName_G = Name_G.drop([0, 1, 2, 3])

Name_H = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_H}.xls')
NewName_H = Name_H.drop([0, 1, 2, 3])

Name_I = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_I}.xls')
NewName_I = Name_I.drop([0, 1, 2, 3])

Name_J = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_J}.xls')
NewName_J = Name_J.drop([0, 1, 2, 3])

Name_K = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_K}.xls')
NewName_K = Name_K.drop([0, 1, 2, 3])

Name_L = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_L}.xls')
NewName_L = Name_L.drop([0, 1, 2, 3])

Name_M = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_M}.xls')
NewName_M = Name_M.drop([0, 1, 2, 3])

Name_N = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_N}.xls')
NewName_N = Name_N.drop([0, 1, 2, 3])

Name_O = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_O}.xls')
NewName_O = Name_O.drop([0, 1, 2, 3])

Name_P = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_P}.xls')
NewName_P = Name_P.drop([0, 1, 2, 3])

Name_Q = pd.read_excel(
    f'C:\\Users\\Administrator\\Downloads\\{Name_Q}.xls')
NewName_Q = Name_Q.drop([0, 1, 2, 3])

Name_concat = pd.concat([NewName_A, NewName_B, NewName_C, NewName_D, NewName_E, NewName_F, NewName_G, NewName_H,
                         NewName_I, NewName_J, NewName_K, NewName_L, NewName_M, NewName_N, NewName_O, NewName_P, NewName_Q], axis=0)
Name_concat.to_excel(
    f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')

# 샘플 3개
# Name_concat = pd.concat([NewName_A, NewName_B, NewName_C], axis=0)
# Name_concat.to_excel(
#     f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')

# 첫 번째 행 수정
wb = openpyxl.load_workbook(
    f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')

ws = wb.active

ws['A1'] = 'No'
ws['B1'] = '점검업체'
ws['C1'] = '점검일자'
ws['D1'] = '시작시간'
ws['E1'] = '종료시간'
ws['F1'] = '전산등록일'
ws['G1'] = '전산수정일'
ws['H1'] = '건물명'
ws['I1'] = '건물주소1'
ws['J1'] = '건물주소2'
ws['K1'] = '승강기고유번호'
ws['L1'] = '호기'
ws['M1'] = '설치장소'
ws['N1'] = '승강기종류'
ws['O1'] = '점검자'
ws['P1'] = '점검보조자'
ws['Q1'] = '관리주체확인여부'
ws['R1'] = '관리주체확인일자'
ws['S1'] = '점검결과'

wb.save(
    f'C:\\Users\\Administrator\\Downloads\\2022년+{Month}월+자체점검실시내역{YMD}_통합.xlsx')
